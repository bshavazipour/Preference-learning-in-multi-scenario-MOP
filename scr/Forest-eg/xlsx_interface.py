#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl as oxl
import os
from shutil import copy as copy_file
from copy import copy
import numpy as np

ref_table_title = "Reference point"
sol_table_title = "Solution"
# Colors of cells with aspiration levels
black_color="00000000"
blue_color="000000FF"
font_set = oxl.styles.Font(color=blue_color) # Set by the DM
font_calc = oxl.styles.Font(color=black_color) # Calculated by the method

# Shortening for get_column_letter function
gcl = oxl.utils.get_column_letter
# Converts row and column nr. to cell coordinates in excel notation (str) 
def cc(row,col):
    return gcl(col)+str(row)

# Given a worksheet, returns (nr. of criteria, nr. of scenarios) based on
#   the shape of the first table containing ideal, nadir information
def get_problem_shape(ws):
    # Nr. of criteria
    n_crit = ws.max_column-2 # 1st guess assuming there is only one table
    # Correction if there are more tables, based on empty column between them
    for i in range(1,ws.max_column+1):
        if ws[cc(2,i)].value is None:
            n_crit=i-3
            break
    # Nr. of scenarios
    n_scen = int((ws.max_row-2)/2)
    return (n_crit, n_scen)

# Given a worksheet and a table title, returns nr. of column with this title
#   * nr. of scenarios is taken from the 1st row
def get_table_col(ws,title):
    col = None
    for i,ci in enumerate(ws[1]):
        if ci.value == title:
            col = i+1
            break
    return col

# Given a worksheet, column nr, and table title, creates an empty table
# by copying headers and styles from the first (ideal,nadir) table
def add_table(ws,col_n,title,font=font_calc):
    n_crit,n_scen = get_problem_shape(ws)
    ws[cc(1,col_n)] = title
    ws[cc(1,col_n)]._style = copy(ws["A1"]._style)
    for i,irow in enumerate(ws.iter_rows(
            min_row=2, max_col=n_crit+2, max_row=2*n_scen+2
            )):
        j=0 # column number; 2nd column is skipped
        for j0,icell in enumerate(irow):
            # Skipping 2nd column
            if j0==1:
                continue
            # Changing column width
            if i==0:
                ws.column_dimensions[gcl(col_n+j)].width =\
                    ws.column_dimensions[gcl(j0+1)].width
            # Copying cell values for header and col. names
            if i==0 or j==0:
                try:
                    ws[cc(i+2,col_n+j)] = icell.value
                except:
                    pass
            # Copying cell format
            try:
                ws[cc(i+2,col_n+j)]._style = copy(icell._style)
            except:
                pass
            # Changing cell color - all cells are DM-sset
            if i>0 and j>0:
                ws[cc(i+2,col_n+j)].font = font
                ws[cc(i+2,col_n+j)].alignment = oxl.styles.Alignment(
                        vertical="center"
                        )
            # Merging cells
            if i/2 != int(i/2):
                ws.merge_cells(None,i+2,col_n+j,i+3,col_n+j)
            j+=1    

# Given a worksheet, a table title, and an array, sets values in the table
# from the array, only for cells with black color
def set_table_values(ws,title,array):
    col0 = get_table_col(ws,title)
    if col0 is None:
        print(10*"-","\n",f"! Table \'{title}\' is not created !\n",10*"-")
        return
    for i, rowi in enumerate(array):
        for j, x in enumerate(rowi):
            cell = ws[cc(3+i*2, col0+j+1)]
            if (not cell.font.color) or cell.font.color.rgb==black_color or\
                cell.value is None:
                cell.value = x
                cell.font = font_calc
            else:
                cell.font = font_set

# Initialize a new interactive process by creating a template xlsx file
# from the file containing problem information
def new_interaction(init_file, new_file):
    wb = oxl.load_workbook(init_file)
    ws = wb[wb.sheetnames[0]]
    n_crit, n_scen = get_problem_shape(ws)
    add_table(ws,ws.max_column + 2,ref_table_title,font=font_set)
    wb.save(new_file)
    wb.close()

# Given xlsx file name, returns a numpy array containing aspiration levels
# set by the DM, and the remaining values NaN. 
#   - NaN is produced by an empty cell, or a cell of black color
def get_preferences(file_name):
    wb = oxl.load_workbook(file_name)
    ws = wb[wb.sheetnames[0]]    
    col0 = get_table_col(ws,ref_table_title)
    if col0 is None:
        print(10*"-","\n",f"! Table \'{ref_table_title}\' is not created !\n",10*"-")
        return
    n_crit, n_scn = get_problem_shape(ws)
    array = np.empty((n_scn,n_crit))
    for i in range(n_scn):
        for j in range(n_crit):
            cell = ws[cc(3+i*2,col0+j+1)]
            if cell.font.color and cell.font.color.rgb!=black_color:
                array[i,j]=cell.value
    wb.close()
    return array

# Given previous and next xlsx file names, estimated preference matrix and
#   optoinally, solution matrix, creates the next file with updated information
#   * in the preference table, only black or empty cells are updated
def next_iteration(file_previous,file_next,pref_array,sol_array=None):
    wb = oxl.load_workbook(file_previous)
    ws = wb[wb.sheetnames[0]]    
    # Preference table
    set_table_values(ws,ref_table_title,pref_array)
    # Solution table
    col0 = get_table_col(ws,sol_table_title)
    if sol_array is None: 
        if col0 is not None: # removing solution table
            n_crit,_ = get_problem_shape(ws)
            ws.delete_cols(col0,n_crit+1)
    else: # adding solution information
        if col0 is None:
            add_table(ws, ws.max_column+2, sol_table_title)
        set_table_values(ws,sol_table_title,sol_array)
        
    wb.save(file_next)
    wb.close()    

if __name__ == "__main__":
    fld = "/home/dp/Dropbox/Scenarios_MOO/Forest problem/"
    # new_interaction(
    #     "/home/dp/Dropbox/Scenarios_MOO/Forest problem/Pareto ranges.xlsx",
    #     "/home/dp/Dropbox/Scenarios_MOO/Forest problem/process")
    # ws = oxl.load_workbook("/home/dp/Dropbox/Scenarios_MOO/Forest problem/Pareto ranges.xlsx")[
    #     "Arkusz1"]
    # print(get_table_col(ws,"Combined for all scenarios"))

    # 1. Create new interaction
    #new_interaction("test/_init_file.xlsx","test/iter_0.xlsx")
    
    # 2. After the DM set partial preferences, get them from the file
    # pref_matrix = get_preferences("test/_iter_0_prefs.xlsx")
    # print(pref_matrix)
    #3. After calculating missing preferences and solution, update information
    np.nan_to_num(pref_matrix,copy=False,nan=100000)
    next_iteration(
        "test/_iter_0_prefs.xlsx", "test/iter_1.xlsx",
        pref_matrix,sol_array=np.full((12,4),200000)
        )